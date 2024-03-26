# -*- coding: utf-8 -*-

import os

import faiss
import pandas as pd
import pdfplumber
import docx2txt
import openpyxl
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import UnstructuredHTMLLoader

from openai_helper import *

text_splitter = RecursiveCharacterTextSplitter.from_tiktoken_encoder(
    chunk_size=800,
    chunk_overlap=40
)


def download_data(url, save_path):
    try:
        response = requests.get(url)
        with open(save_path, 'wb') as file:
            file.write(response.content)
        return save_path
    except:
        logger.exception("download_data {} failed...".format(url))


def process_eml_his_data(save_path):
    try:
        bot_path = "/".join(save_path.split("/")[:-1])
        email_history_emb_faiss = os.path.join(bot_path, EMAIL_HISTORY_EMB + ".faiss")
        email_history_emb_txt = os.path.join(bot_path, EMAIL_HISTORY_EMB + ".txt")

        content4embs, email_histories = [], []
        with open(save_path, "r") as file:
            for line in file:
                json_obj = json.loads(line)
                content4emb, email_history = extract_email_history(json_obj)
                content4embs.append("\n".join(content4emb))
                email_histories.append("\n".join(email_history))

        content_embs = get_batch_emb(content4embs)
        df_eml_his_emb = pd.DataFrame({"content4emb": content4embs, "content": email_histories})
        df_eml_his_emb.drop_duplicates(subset='content4emb', inplace=True)

        df_eml_his_emb.to_csv(email_history_emb_txt, index=False, encoding="utf-8")
        # faiss index
        index = faiss.IndexFlatIP(1536)
        embs = np.array(content_embs).astype('float32')
        index.add(embs)
        faiss.write_index(index, email_history_emb_faiss)

        logger.info("{} eml_his_emb successfully---{}".format(bot_path, df_eml_his_emb))
        return df_eml_his_emb
    except:
        logger.exception("process_eml_his_data failed...")


def process_chat_his_data(save_path):
    try:
        bot_path = "/".join(save_path.split("/")[:-1])
        chat_history_emb_faiss = os.path.join(bot_path, CHAT_HISTORY_EMB + ".faiss")
        chat_history_emb_txt = os.path.join(bot_path, CHAT_HISTORY_EMB + ".txt")

        content4embs, chat_histories = [], []
        with open(save_path, "r") as file:
            for line in file:
                json_obj = json.loads(line)
                content4emb, chat_history = extract_chat_history(json_obj)
                content4embs.append(content4emb)
                chat_histories.append(chat_history)

        content_embs = get_batch_emb(content4embs)
        df_chat_his_emb = pd.DataFrame({"content4emb": content4embs, "content": chat_histories})
        df_chat_his_emb.drop_duplicates(subset='content4emb', inplace=True)

        df_chat_his_emb.to_csv(chat_history_emb_txt, index=False, encoding="utf-8")
        # faiss index
        index = faiss.IndexFlatIP(1536)
        embs = np.array(content_embs).astype('float32')
        index.add(embs)
        faiss.write_index(index, chat_history_emb_faiss)

        logger.info("{} chat_his_emb successfully---{}".format(bot_path, df_chat_his_emb))
        return df_chat_his_emb
    except:
        logger.exception("process_chat_his_data failed...")


def process_bot_data(save_path):
    try:
        bot_path = "/".join(save_path.split("/")[:-1])
        bot_emb_faiss = os.path.join(bot_path, BOT_EMB + ".faiss")
        bot_emb_txt = os.path.join(bot_path, BOT_EMB + ".txt")

        content4embs, qas = [], []
        with open(save_path, "r") as file:
            for line in file:
                json_obj = json.loads(line)
                content4emb, qa = extract_qa(json_obj)
                content4embs.extend(content4emb)
                qas.extend(qa)

        content_embs = get_batch_emb(content4embs)
        df_bot_emb = pd.DataFrame({"content4emb": content4embs, "content": qas})
        df_bot_emb.drop_duplicates(subset='content4emb', inplace=True)

        df_bot_emb.to_csv(bot_emb_txt, index=False, encoding="utf-8")
        # faiss index
        index = faiss.IndexFlatIP(1536)
        embs = np.array(content_embs).astype('float32')
        index.add(embs)
        faiss.write_index(index, bot_emb_faiss)

        logger.info("{} bot_emb successfully---{}".format(bot_path, df_bot_emb))
        return df_bot_emb
    except:
        logger.exception("process_bot_data failed...")


def process_kb_data(save_path):
    try:
        bot_path = "/".join(save_path.split("/")[:-1])
        kb_emb_faiss = os.path.join(bot_path, KB_EMB + ".faiss")
        kb_emb_txt = os.path.join(bot_path, KB_EMB + ".txt")

        content4embs, qas = [], []
        with open(save_path, "r") as file:
            for line in file:
                json_obj = json.loads(line)
                content4emb, qa = extract_kb(json_obj)
                content4embs.extend(content4emb)
                qas.extend(qa)

        content_embs = get_batch_emb(content4embs)
        df_kb_emb = pd.DataFrame({"content4emb": content4embs, "content": qas})
        df_kb_emb.drop_duplicates(subset='content4emb', inplace=True)

        df_kb_emb.to_csv(kb_emb_txt, index=False, encoding="utf-8")
        # faiss index
        index = faiss.IndexFlatIP(1536)
        embs = np.array(content_embs).astype('float32')
        index.add(embs)
        faiss.write_index(index, kb_emb_faiss)

        logger.info("{} kb_emb successfully---{}".format(bot_path, df_kb_emb))
        return df_kb_emb
    except:
        logger.exception("process_kb_data failed...")


def download_extract_data(data_urls, bot_path):
    file_content_dic = {}
    for data_url in data_urls:
        try:
            response = requests.get(data_url.strip())
            file_name = data_url.split("/")[-1]
            tmp_file_path = os.path.join(bot_path, file_name)
            with open(tmp_file_path, 'wb') as file:
                file.write(response.content)

            # 文件解析
            if tmp_file_path.endswith(".pdf"):
                with pdfplumber.open(tmp_file_path) as pdf:
                    text = "".join([page.extract_text() for page in pdf.pages])
            elif tmp_file_path.endswith(".docx") or tmp_file_path.endswith(".doc"):
                text = docx2txt.process(tmp_file_path)
            elif tmp_file_path.endswith(".xlsx") or tmp_file_path.endswith(".xls"):
                wookbook = openpyxl.load_workbook(tmp_file_path)
                sns = wookbook.sheetnames
                all_contents = []
                for sn in sns:
                    worksheet = wookbook[sn]
                    sheet_contents = worksheet.title + "\n"

                    mr = worksheet.max_row
                    cols = list(worksheet.iter_cols())
                    for i in range(mr):
                        sheet_contents += "\t".join([str(col[i].value) for col in cols]) + "\n"

                    all_contents.append(sheet_contents)
                text = "\n".join(all_contents)
            else:
                with open(tmp_file_path, "r") as file:
                    text = file.read()

            file_content_dic[file_name] = text
        except:
            logger.exception("download_extract_data {} failed...".format(data_url))
    return file_content_dic


def process_custom_data(file_content_dic, bot_path):
    try:
        custom_emb_faiss = os.path.join(bot_path, CUSTOM_EMB + ".faiss")
        custom_emb_txt = os.path.join(bot_path, CUSTOM_EMB + ".txt")

        content4embs = []
        file_names = []
        for file_name, file_content in file_content_dic.items():
            docs = text_splitter.split_text(file_content)
            for doc in docs:
                content4embs.append(doc)
                file_names.append(file_name)

        content_embs = get_batch_emb(content4embs)
        df_custom_emb = pd.DataFrame({"content4emb": content4embs, "src": file_names})
        df_custom_emb.drop_duplicates(subset='content4emb', inplace=True)

        df_custom_emb.to_csv(custom_emb_txt, index=False, encoding="utf-8")
        # faiss index
        index = faiss.IndexFlatIP(1536)
        embs = np.array(content_embs).astype('float32')
        index.add(embs)
        faiss.write_index(index, custom_emb_faiss)

        logger.info("{} custom_emb successfully---{}".format(bot_path, df_custom_emb))
        return df_custom_emb
    except:
        logger.exception("process_custom_data failed...")


def process_spider_data(file_content_dic, bot_path):
    try:
        spider_emb_faiss = os.path.join(bot_path, SPIDER_EMB + ".faiss")
        spider_emb_txt = os.path.join(bot_path, SPIDER_EMB + ".txt")

        content4embs = []
        file_names = []
        for file_name, file_content in file_content_dic.items():
            docs = text_splitter.split_text(file_content)
            for doc in docs:
                content4embs.append(doc)
                file_names.append(file_name)

        content_embs = get_batch_emb(content4embs)
        df_spider_emb = pd.DataFrame({"content4emb": content4embs, "src": file_names})
        df_spider_emb.drop_duplicates(subset='content4emb', inplace=True)

        df_spider_emb.to_csv(spider_emb_txt, index=False, encoding="utf-8")
        # faiss index
        index = faiss.IndexFlatIP(1536)
        embs = np.array(content_embs).astype('float32')
        index.add(embs)
        faiss.write_index(index, spider_emb_faiss)

        logger.info("{} spider_url_emb successfully---{}".format(bot_path, df_spider_emb))
        return df_spider_emb
    except:
        logger.exception("process_spider_data failed...")


def extract_query(json_obj):
    content4emb, email_history = [], []
    content4emb.append(json_obj["subject"])

    if_chat = all(":" in element for element in json_obj["history"])

    for email_content in json_obj["history"]:
        if if_chat:
            colon_first_idx = email_content.find(":")
            if email_content[:colon_first_idx].strip().lower() == "user":
                content4emb.append(email_content[colon_first_idx + 1:].strip())
            email_history.append(email_content.strip())
        else:
            content4emb.append(email_content.strip())
            email_history.append("Question:" + email_content.strip())

    # 取最新轮
    if len(content4emb) > 1:
        if len(content4emb[-1].strip().split()) <= 4:
            content4emb = content4emb[-2:]
        else:
            content4emb = content4emb[-1:]
    return content4emb, email_history


def extract_html(src_dir):
    file_content_dic = {}
    for file_name in os.listdir(src_dir):
        # with open(os.path.join(src_dir, file_name), 'r') as file:
        #     f_content = html2text.html2text(file.read()).strip()
        loader = UnstructuredHTMLLoader(os.path.join(src_dir, file_name))
        documents = loader.load()
        f_content = "\n".join([document.page_content for document in documents])
        if file_name.endswith(".html"):
            file_name_cont = ".".join(file_name.split(".")[:-1]).strip()
        else:
            file_name_cont = file_name
        file_content_dic[file_name] = file_name_cont + "\n" + f_content
    return file_content_dic


def extract_email_history(json_obj):
    content4emb, email_history = [], []
    content4emb.append(json_obj["subject"])

    for email_content in json_obj["history"]:
        text = email_content.split("Subject:")[1].strip()
        content4emb.append(text)

        person = email_content.split("\n")[0].replace("From:", "").strip()
        email_history.append(person + ":" + text)
    return content4emb, email_history


def extract_chat_history(json_obj):
    content4emb, chat_history = [], []
    chat_his_all_content = json_obj["content"]
    messages = chat_his_all_content.split("\n")
    for i in range(len(messages) - 1, -1, -1):
        if bool(re.match(r'^\[\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}]:', messages[i])):
            mess_splits = messages[i].split(":")
            content4emb.insert(0, ":".join(mess_splits[4:]))
            chat_history.insert(0, ":".join(mess_splits[3:]))
        else:
            messages[i - 1] = messages[i - 1] + "\\n" + messages[i]
    return "\n".join(content4emb), "\n".join(chat_history)


def extract_qa(json_obj):
    content4emb, qa = [], []
    qs = json_obj["questions"]
    qs.append(json_obj["intentName"])

    a = json_obj["answer"]
    for q in qs:
        content4emb.append(q)
        q_a = "Question:" + q + "\nAnswer:" + a
        qa.append(q_a)
    return content4emb, qa


def extract_kb(json_obj):
    content4emb, qa = [], []
    qs = json_obj["similarQuestions"]
    qs.append(json_obj["title"])

    a = json_obj["content"]
    for q in qs:
        content4emb.append(q)
        q_a = "Question:" + q + "\nAnswer:" + a
        qa.append(q_a)
    return content4emb, qa


def filter_empty_mess(messages):
    filtered_messes = []
    for m in messages:
        if m.strip() == "" or m.split(":")[1].strip() == "":
            continue
        else:
            filtered_messes.append(m)
    return filtered_messes


def truncate_massages(messages, token_limit):
    idx_final = []
    messages_final = []
    cur_tok_cnt = 0
    cur_i = 0
    for i in range(len(messages) - 1, -1, -1):
        cur_tok_cnt += count_token(messages[i])
        if cur_tok_cnt > token_limit:
            break
        messages_final.insert(0, messages[i])
        idx_final.insert(0, i)
        cur_i = i
    return messages_final, messages[:cur_i], idx_final
